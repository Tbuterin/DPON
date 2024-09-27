import openpyxl
from openpyxl import Workbook, load_workbook

class ExcelHandler:
    def __init__(self, filename, add_info=None):
        self.filename = filename
        if add_info:
            self.filename = self.filename[:-5] + add_info + ".xlsx"
        self.workbook = self._create_or_load_workbook()

    def _create_or_load_workbook(self):
        try:
            # 尝试加载已经存在的工作簿
            workbook = load_workbook(self.filename)
            # print(f"Loaded existing workbook: {self.filename}")
        except FileNotFoundError:
            # 如果文件不存在，则创建一个新的工作簿
            workbook = Workbook()
            # 删除默认创建的Sheet
            # workbook.remove(workbook.active)
            # workbook.save(self.filename)
            # print(f"Created new workbook: {self.filename}")
        return workbook

    def add(self, data, sheet_name='Sheet'):
        # 获取指定的工作表，如果不存在则创建一个新的工作表
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.create_sheet(sheet_name)
            # print(f"Created new sheet: {sheet_name}")

        # 找到第一行空行
        row = sheet.max_row + 1

        # 将数据写入工作表
        for col, value in enumerate(data, start=1):
            sheet.cell(row=row, column=col, value=value)
        # print(f"Added data to {sheet_name} at row {row}: {data}")

    def save(self):
        # 保存工作簿
        self.workbook.save(self.filename)
        # print(f"Workbook saved: {self.filename}")